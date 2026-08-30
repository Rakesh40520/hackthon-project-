"""XLSX extraction."""
from __future__ import annotations

from typing import List

from app.document_processing.extractor import ExtractedDocument, ExtractedSection


def extract_xlsx(file_path: str, filename: str) -> ExtractedDocument:
    from openpyxl import load_workbook

    wb = load_workbook(file_path, data_only=True, read_only=True)
    sections: List[ExtractedSection] = []
    tables: List[List[List[str]]] = []
    full_text_parts: List[str] = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        grid: List[List[str]] = []
        for row in ws.iter_rows(values_only=True):
            line = [str(c) if c is not None else "" for c in row]
            if any(c.strip() for c in line):
                grid.append(line)
                full_text_parts.append(" | ".join(line))
        if grid:
            tables.append(grid)
            sections.append(
                ExtractedSection(
                    text="\n".join(" | ".join(r) for r in grid),
                    section=str(sheet_name),
                    document=filename,
                )
            )
    wb.close()
    return ExtractedDocument(
        text="\n".join(full_text_parts),
        sections=sections,
        tables=tables,
        metadata={"filename": filename, "type": "xlsx"},
    )
