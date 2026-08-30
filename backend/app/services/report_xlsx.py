"""Excel report generation."""
from __future__ import annotations

import io
from datetime import datetime
from typing import Dict, List

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from app.models import ProcurementProject, Proposal, Requirement, RequirementEvaluation


def _header(ws, headers, row=1):
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")


def _vendor_name(p: Proposal) -> str:
    return p.project_vendor.vendor.company_name if p.project_vendor and p.project_vendor.vendor else "—"


def generate_xlsx_report(
    project: ProcurementProject,
    proposals: List[Proposal],
    requirements: List[Requirement],
    evaluations: Dict[str, List[RequirementEvaluation]],
) -> bytes:
    wb = Workbook()
    sorted_props = sorted(
        [p for p in proposals if p.score],
        key=lambda p: (not p.score.is_eligible, -(p.score.total_score or 0)),
    )

    # Sheet 1: Executive Summary
    ws = wb.active
    ws.title = "Executive Summary"
    ws["A1"] = f"Procurement Report: {project.name}"
    ws["A1"].font = Font(size=16, bold=True)
    ws["A2"] = f"Generated: {datetime.utcnow().isoformat()}"
    rows = [
        ("Project", project.name),
        ("Category", project.category),
        ("Budget", f"{project.currency} {project.budget or ''}"),
        ("Status", project.status.value),
        ("Proposals", len(proposals)),
        ("Requirements", len(requirements)),
    ]
    for i, (k, v) in enumerate(rows, start=4):
        ws.cell(row=i, column=1, value=k).font = Font(bold=True)
        ws.cell(row=i, column=2, value=v)

    # Sheet 2: Ranking
    ws2 = wb.create_sheet("Ranking")
    _header(ws2, ["Rank", "Vendor", "Score", "Eligibility", "Year 1", "Year 3", "Year 5"])
    for i, p in enumerate(sorted_props, start=1):
        ws2.append([
            i, _vendor_name(p), round(p.score.total_score, 1),
            "Eligible" if p.score.is_eligible else "Ineligible",
            p.pricing.year1_total if p.pricing and p.pricing.year1_total else None,
            p.pricing.year3_total if p.pricing and p.pricing.year3_total else None,
            p.pricing.year5_total if p.pricing and p.pricing.year5_total else None,
        ])

    # Sheet 3: Requirements matrix
    ws3 = wb.create_sheet("Requirements")
    _header(ws3, ["Requirement"] + [_vendor_name(p) for p in sorted_props])
    for r in requirements:
        row = [r.name]
        for p in sorted_props:
            evals = evaluations.get(str(p.id), [])
            ev = next((e for e in evals if e.requirement_id == r.id), None)
            row.append(ev.status.value if ev else "—")
        ws3.append(row)

    # Sheet 4: Risks
    ws4 = wb.create_sheet("Risks")
    _header(ws4, ["Vendor", "Title", "Category", "Severity", "Description", "Recommendation"])
    for p in sorted_props:
        for r in p.risks:
            ws4.append([_vendor_name(p), r.title, r.category.value, r.severity.value, r.description, r.recommendation])

    # Sheet 5: AI Recommendation
    ws5 = wb.create_sheet("AI Recommendation")
    _header(ws5, ["Vendor", "Decision", "Summary", "Reasoning", "Strengths", "Weaknesses", "Next Steps"])
    for p in sorted_props:
        if p.recommendation:
            ws5.append([
                _vendor_name(p), p.recommendation.decision,
                p.recommendation.summary, p.recommendation.reasoning,
                "; ".join(p.recommendation.strengths or []),
                "; ".join(p.recommendation.weaknesses or []),
                "; ".join(p.recommendation.next_steps or []),
            ])

    # Sheet 6: Missing Information
    ws6 = wb.create_sheet("Missing Information")
    _header(ws6, ["Vendor", "Field", "Importance", "Why it matters"])
    for p in sorted_props:
        for m in p.missing_info:
            ws6.append([_vendor_name(p), m.field_name, m.importance, m.why_it_matters])

    # Sheet 7: Clarification Questions
    ws7 = wb.create_sheet("Clarification Questions")
    _header(ws7, ["Vendor", "Question", "Category", "Priority"])
    for p in sorted_props:
        for c in p.clarification_questions if hasattr(p, "clarification_questions") else []:
            ws7.append([_vendor_name(p), c.question, c.category, c.priority])

    # Auto width
    for sheet in wb.worksheets:
        for column_cells in sheet.columns:
            try:
                length = max(len(str(c.value or "")) for c in column_cells)
            except Exception:
                length = 10
            sheet.column_dimensions[column_cells[0].column_letter].width = min(60, length + 2)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
